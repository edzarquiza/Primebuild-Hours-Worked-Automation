import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import timedelta, datetime
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Primebuild Hours Worked Automation",
    page_icon="🏗️",
    layout="wide",
)

# ── Styles ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f4f6f9; }
.main-header {
    background: linear-gradient(135deg, #1a3a5c 0%, #2d6a9f 100%);
    color: white; padding: 2rem 2.5rem; border-radius: 12px;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 15px rgba(26,58,92,0.3);
}
.main-header h1 { margin: 0; font-size: 2rem; font-weight: 700; }
.main-header p  { margin: .4rem 0 0; opacity: .85; font-size: 1rem; }
.metric-card {
    background: white; border-radius: 10px; padding: 1.2rem 1.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,.08); text-align: center;
}
.metric-card .num  { font-size: 2.2rem; font-weight: 700; color: #1a3a5c; }
.metric-card .lbl  { font-size: .85rem; color: #666; margin-top: .2rem; }
.metric-card.warn  .num { color: #e67e22; }
.metric-card.danger .num { color: #e74c3c; }
.result-box {
    background: white; border-radius: 10px; padding: 1.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 1rem;
}
.badge {
    display:inline-block; padding:.2rem .7rem; border-radius:20px;
    font-size:.8rem; font-weight:600;
}
.badge-yellow { background:#fff3cd; color:#856404; }
.badge-red    { background:#f8d7da; color:#842029; }
.badge-green  { background:#d1e7dd; color:#0f5132; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>🏗️ Primebuild Hours Worked Automation</h1>
  <p>Upload one or more weekly timesheet exports — compliance reports are generated instantly.</p>
</div>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
LONG_SHIFT_THRESHOLD   = 14.0   # hours
SHORT_BREAK_THRESHOLD  = 10.0   # hours
WEEKLY_HOURS_THRESHOLD = 60.0   # hours
FATIGUE_SUM_THRESHOLD  = 14.0   # combined hours triggering fatigue flag

YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
RED_FILL    = PatternFill("solid", fgColor="FF0000")

# ── Helper: parse duration strings like "8:30:00" → float hours ──────────────
def parse_duration(val) -> float:
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    m = re.match(r'^(\d+):(\d+):(\d+)$', s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60 + int(m.group(3)) / 3600
    try:
        return float(s)
    except Exception:
        return 0.0

def timedelta_to_hours(td) -> float:
    if isinstance(td, pd.Timedelta):
        return td.total_seconds() / 3600
    return parse_duration(td)

# ── Core processing ───────────────────────────────────────────────────────────
def process_file(uploaded_file) -> dict:
    """
    Returns a dict with keys:
      raw_df, long_shift_df, weekly_df, summary, filename_stem
    """
    xl = pd.ExcelFile(uploaded_file)
    df = pd.read_excel(xl, sheet_name="Export", header=0)

    stem = re.sub(r'\.xlsx?$', '', uploaded_file.name, flags=re.I)

    # ── 1. Filter: keep Shift work + blank Work Type rows (mirrors VBA logic)
    #    The VBA deletes rows where Work Type is NOT "Shift work" AND NOT blank
    #    Also exclude any "Totals" summary rows (col A = "Totals")
    wt = df["Work Type"].astype(str).str.strip()
    df_shift = df[
        ((wt.str.lower() == "shift work") |
        (df["Work Type"].isna()) |
        (wt == "")) &
        (df["Employee Id"].astype(str).str.strip().str.lower() != "totals")
    ].copy()

    if df_shift.empty:
        return {
            "raw_df": df, "long_shift_df": pd.DataFrame(),
            "weekly_df": pd.DataFrame(), "summary": {},
            "filename_stem": stem, "error": "No shift data rows found."
        }

    # ── 2. Build datetime columns ─────────────────────────────────────────────
    def make_dt(date_col, time_col):
        dates = pd.to_datetime(df_shift[date_col], errors="coerce")
        times = df_shift[time_col].apply(
            lambda v: v if isinstance(v, pd.Timedelta) else pd.to_timedelta(str(v), errors="coerce")
        )
        return dates + times

    df_shift["shift_start"] = make_dt("Start Date", "Start Time")
    df_shift["shift_end"]   = make_dt("End Date",   "End Time")

    # Duration in hours
    df_shift["duration_hrs"] = df_shift["Duration"].apply(parse_duration)

    # ── 3. Sort: employee → date → time ──────────────────────────────────────
    df_shift["full_name"] = (
        df_shift["First Name"].astype(str).str.strip() + " " +
        df_shift["Surname"].astype(str).str.strip()
    )
    df_shift.sort_values(["full_name", "shift_start"], inplace=True)
    df_shift.reset_index(drop=True, inplace=True)

    # ── 4. Long shift detection (>14 h) ──────────────────────────────────────
    df_shift["long_shift_flag"] = df_shift["duration_hrs"] > LONG_SHIFT_THRESHOLD

    # ── 5. Break time calculation ─────────────────────────────────────────────
    df_shift["break_before_hrs"] = np.nan
    df_shift["short_break_flag"] = False

    for i in range(1, len(df_shift)):
        prev = df_shift.iloc[i - 1]
        curr = df_shift.iloc[i]
        if prev["full_name"] == curr["full_name"]:
            gap = (curr["shift_start"] - prev["shift_end"]).total_seconds() / 3600
            df_shift.at[df_shift.index[i], "break_before_hrs"] = round(gap, 2)
            if gap < SHORT_BREAK_THRESHOLD:
                df_shift.at[df_shift.index[i], "short_break_flag"] = True

    # ── 6. Fatigue risk: short break AND combined hours (this + prev) > 14h ─────
    # Mirrors VBA: SumDuration = R(x) + R(x-1); if > 0.584 (14h) → highlight
    df_shift["fatigue_risk_flag"] = False

    for i in range(1, len(df_shift)):
        if not df_shift.at[df_shift.index[i], "short_break_flag"]:
            continue
        prev = df_shift.iloc[i - 1]
        curr = df_shift.iloc[i]
        if prev["full_name"] != curr["full_name"]:
            continue
        combined = (
            df_shift.at[df_shift.index[i - 1], "duration_hrs"] +
            df_shift.at[df_shift.index[i],     "duration_hrs"]
        )
        if combined > FATIGUE_SUM_THRESHOLD:
            df_shift.at[df_shift.index[i - 1], "fatigue_risk_flag"] = True
            df_shift.at[df_shift.index[i],     "fatigue_risk_flag"] = True

    # ── 7. Long shift report — mirrors VBA: only rows where Duration is yellow ──
    # Yellow = long shift (>14h) OR fatigue risk (short break + combined > 14h)
    # Short break alone (combined ≤ 14h) is NOT included
    long_shift_df = df_shift[
        df_shift["long_shift_flag"] |
        df_shift["fatigue_risk_flag"]
    ].copy()
    long_shift_df = long_shift_df.sort_values(
        ["Employee Id", "shift_start"]
    ).reset_index(drop=True)

    # ── 8. Weekly hours summary ───────────────────────────────────────────────
    weekly = (
        df_shift.groupby("Employee Id")["duration_hrs"]
        .sum()
        .reset_index()
        .rename(columns={"Employee Id": "Employee", "duration_hrs": "Total Hours"})
    )
    weekly["Total Hours"] = weekly["Total Hours"].round(2)
    weekly["Exceeds 60h"] = weekly["Total Hours"] > WEEKLY_HOURS_THRESHOLD
    weekly.sort_values("Total Hours", ascending=False, inplace=True)

    summary = {
        "total_employees":  df_shift["full_name"].nunique(),
        "total_shifts":     len(df_shift),
        "long_shifts":      int(df_shift["long_shift_flag"].sum()),
        "short_breaks":     int(df_shift["fatigue_risk_flag"].sum()),
        "fatigue_flags":    int(df_shift["fatigue_risk_flag"].sum()),
        "exceed_60h":       int(weekly["Exceeds 60h"].sum()),
    }

    return {
        "raw_df": df,
        "shift_df": df_shift,
        "long_shift_df": long_shift_df,
        "weekly_df": weekly,
        "summary": summary,
        "filename_stem": stem,
    }

# ── 29-column layout matching the VBA output exactly ─────────────────────────
# Dimension (orig col 8) and Reviewed Date UTC (orig col 30) are dropped.
# A blank col L is inserted between End Time (col 11) and Actual Start Date (col 13).
HEADERS_29 = [
    "Employee Id", "First Name", "Surname", "Employee External Id",
    "Timesheet Id", "Status", "Location",
    "Start Date", "Start Time", "End Date", "End Time",
    None,                          # col 12 — inserted break flag, no header
    "Actual Start Date", "Actual Start Time", "Actual End Date", "Actual End Time",
    "Time Variance", "Duration", "Total Duration", "Units", "Unit Type",
    "Work Type", "Shift Conditions", "Classification",
    "Number Of Breaks", "Break Duration",
    "Consolidated With Timesheet Line Id", "Reviewed By", "Created Date UTC",
]  # 29 entries — Duration is at position 18 (col R), matching the template


def _raw_row_values(row: pd.Series) -> list:
    """Return 29 values matching HEADERS_29 for one data row."""
    src = row

    def safe_date(v):
        try:
            return pd.to_datetime(v).date() if pd.notna(v) else None
        except Exception:
            return None

    def safe_td(v):
        """Return a timedelta so openpyxl writes it as [h]:mm:ss duration."""
        if isinstance(v, pd.Timedelta):
            return timedelta(seconds=int(v.total_seconds()))
        if isinstance(v, timedelta):
            return v
        try:
            return timedelta(seconds=int(pd.to_timedelta(str(v)).total_seconds()))
        except Exception:
            return None

    return [
        src.get("Employee Id"),                              # 1
        src.get("First Name"),                               # 2
        src.get("Surname"),                                  # 3
        src.get("Employee External Id"),                     # 4
        src.get("Timesheet Id"),                             # 5
        src.get("Status"),                                   # 6
        src.get("Location"),                                 # 7
        # Dimension dropped
        safe_date(src.get("Start Date")),                    # 8
        safe_td(src.get("Start Time")),                      # 9
        safe_date(src.get("End Date")),                      # 10
        safe_td(src.get("End Time")),                        # 11
        "Less than 10 hour break" if row.get("short_break_flag") else "",  # 12 col L
        safe_date(src.get("Actual Start Date")),             # 13
        safe_td(src.get("Actual Start Time")),               # 14
        safe_date(src.get("Actual End Date")),               # 15
        safe_td(src.get("Actual End Time")),                 # 16
        src.get("Time Variance"),                            # 17
        safe_td(src.get("Duration")),                        # 18 col R ← subtotalled
        safe_td(src.get("Total Duration")),                  # 19
        src.get("Units"),                                    # 20
        src.get("Unit Type"),                                # 21
        src.get("Work Type"),                                # 22
        src.get("Shift Conditions"),                         # 23
        src.get("Classification"),                           # 24
        src.get("Number Of Breaks"),                         # 25
        safe_td(src.get("Break Duration")),                  # 26
        src.get("Consolidated With Timesheet Line Id"),      # 27
        src.get("Reviewed By"),                              # 28
        src.get("Created Date UTC"),                         # 29
        # Reviewed Date UTC dropped
    ]


def build_excel(result: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    shift_df  = result["shift_df"]
    long_df   = result["long_shift_df"]

    bold_font   = Font(bold=True, name="Calibri", size=11)
    normal_font = Font(name="Calibri", size=11)

    # ── Sheet 1: Long Shift ───────────────────────────────────────────────────
    ws1 = wb.create_sheet("Long Shift")

    # Write header row
    for ci, h in enumerate(HEADERS_29, 1):
        c = ws1.cell(1, ci, h if h else "")
        c.font = bold_font

    # Write flagged rows — yellow on Duration (col 18) and col L (col 12) when flagged
    src_rows = long_df if not long_df.empty else pd.DataFrame()
    for ri, (_, row) in enumerate(src_rows.iterrows(), 2):
        vals = _raw_row_values(row)
        for ci, v in enumerate(vals, 1):
            ws1.cell(ri, ci, v if v is not None else "").font = normal_font
        # Yellow on Duration (col 18) for every flagged row
        ws1.cell(ri, 18).fill = YELLOW_FILL
        # Yellow on inserted col L (col 12) when short break flag
        if row.get("short_break_flag"):
            ws1.cell(ri, 12).fill = YELLOW_FILL

    # ── Sheet 2: Weekly Hours ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Weekly Hours")

    # Sort by Employee Id then shift time — matches VBA
    wh_df = shift_df.copy()
    wh_df.sort_values(["Employee Id", "shift_start"], inplace=True)
    wh_df.reset_index(drop=True, inplace=True)

    # Write header row (all 29 cols, bold)
    for ci, h in enumerate(HEADERS_29, 1):
        ws2.cell(1, ci, h if h else "").font = bold_font

    # Pre-compute total hours per Employee Id to know who exceeds 60h
    emp_totals = wh_df.groupby("Employee Id")["duration_hrs"].sum()
    exceed_ids = set(emp_totals[emp_totals > WEEKLY_HOURS_THRESHOLD].index.tolist())

    red_font   = Font(bold=True, name="Calibri", size=11, color="FFFFFF")

    # Write data rows grouped by Employee Id with subtotals
    current_row = 2
    grand_total_secs = 0

    for emp_id, grp in wh_df.groupby("Employee Id", sort=False):
        exceeds = emp_id in exceed_ids
        group_secs = int(grp["duration_hrs"].sum() * 3600)
        grand_total_secs += group_secs

        for _, row in grp.iterrows():
            vals = _raw_row_values(row)
            for ci, v in enumerate(vals, 1):
                ws2.cell(current_row, ci, v if v is not None else "").font = normal_font
            ws2.row_dimensions[current_row].outline_level = 2
            # Only expand detail rows for >60h employees
            ws2.row_dimensions[current_row].hidden = not exceeds
            current_row += 1

        # Subtotal row — write actual timedelta value (not formula) so fill works
        subtotal_row = current_row
        c_label = ws2.cell(subtotal_row, 1, f"{emp_id} Total")
        c_dur   = ws2.cell(subtotal_row, 18, timedelta(seconds=group_secs))
        c_dur.number_format = "[h]:mm:ss"
        if exceeds:
            c_label.fill = RED_FILL
            c_dur.fill   = RED_FILL
            c_label.font = red_font
            c_dur.font   = red_font
        else:
            c_label.font = bold_font
            c_dur.font   = bold_font
        ws2.row_dimensions[subtotal_row].outline_level = 1
        ws2.row_dimensions[subtotal_row].hidden = False
        current_row += 1

    # Grand total row — also red if overall total > 60h
    grand_row = current_row
    c_grand     = ws2.cell(grand_row, 1, "Grand Total")
    c_grand_dur = ws2.cell(grand_row, 18, timedelta(seconds=grand_total_secs))
    c_grand_dur.number_format = "[h]:mm:ss"
    if grand_total_secs / 3600 > WEEKLY_HOURS_THRESHOLD:
        c_grand.fill     = RED_FILL
        c_grand_dur.fill = RED_FILL
        c_grand.font     = red_font
        c_grand_dur.font = red_font
    else:
        c_grand.font     = bold_font
        c_grand_dur.font = bold_font
    ws2.row_dimensions[grand_row].hidden = False

    ws2.sheet_properties.outlinePr.summaryBelow = True

    # ── Sheet 3 (Sheet1 placeholder — matches template) ───────────────────────
    wb.create_sheet("Sheet1")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Streamlit UI ──────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload Timesheet Report(s)",
    type=["xlsx"],
    accept_multiple_files=True,
    help="Upload one or more weekly timesheet export files (Excel .xlsx).",
)

if not uploaded:
    st.info("👆 Upload one or more timesheet Excel files to get started.")
    with st.expander("ℹ️ How it works"):
        st.markdown("""
**This tool automatically:**
1. Filters for **Shift Work** entries only
2. Detects **long shifts** exceeding 14 hours
3. Calculates **break gaps** between consecutive shifts and flags breaks under 10 hours
4. Identifies **fatigue risk** — clusters of shifts with short breaks and combined hours > 14h
5. Aggregates **weekly hours** per employee and flags anyone exceeding 60 hours

**Output:** One Excel file per upload with three sheets — *Long Shift*, *Weekly Hours*, and *Summary*.
        """)
    st.stop()

# ── Process each file ─────────────────────────────────────────────────────────
results = []
for f in uploaded:
    with st.spinner(f"Processing {f.name}…"):
        try:
            r = process_file(f)
            r["excel_bytes"] = build_excel(r)
            results.append(r)
        except Exception as e:
            st.error(f"❌ Error processing **{f.name}**: {e}")

if not results:
    st.stop()

# ── Bulk download ─────────────────────────────────────────────────────────────
if len(results) > 1:
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            fname = f"Compliance_{r['filename_stem']}.xlsx"
            zf.writestr(fname, r["excel_bytes"])
    zip_buf.seek(0)
    st.download_button(
        "⬇️ Download All Reports (ZIP)",
        data=zip_buf,
        file_name=f"PrimebuildCompliance_{datetime.now().strftime('%Y%m%d%H%M%S')}.zip",
        mime="application/zip",
        use_container_width=True,
    )
    st.divider()

# ── Per-file results ──────────────────────────────────────────────────────────
for r in results:
    s = r["summary"]
    stem = r["filename_stem"]

    st.markdown(f"### 📄 {stem}")

    if "error" in r:
        st.warning(r["error"])
        continue

    # Metric cards
    cols = st.columns(6)
    cards = [
        ("Employees",       s["total_employees"], ""),
        ("Shifts",          s["total_shifts"],    ""),
        ("Long Shifts",     s["long_shifts"],     "warn" if s["long_shifts"] else ""),
        ("Short Breaks",    s["short_breaks"],    "warn" if s["short_breaks"] else ""),
        ("Fatigue Flags",   s["fatigue_flags"],   "danger" if s["fatigue_flags"] else ""),
        (">60h Employees",  s["exceed_60h"],      "danger" if s["exceed_60h"] else ""),
    ]
    for col, (lbl, num, cls) in zip(cols, cards):
        col.markdown(
            f'<div class="metric-card {cls}"><div class="num">{num}</div>'
            f'<div class="lbl">{lbl}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Preview tabs
    tab1, tab2 = st.tabs(["🚨 Compliance Issues", "📊 Weekly Hours"])

    with tab1:
        ldf = r["long_shift_df"]
        if ldf.empty:
            st.success("✅ No compliance issues detected for this file.")
        else:
            disp = ldf[[
                "full_name", "shift_start", "shift_end",
                "duration_hrs", "break_before_hrs",
                "long_shift_flag", "short_break_flag", "fatigue_risk_flag"
            ]].copy()
            disp.columns = [
                "Employee", "Shift Start", "Shift End",
                "Duration (hrs)", "Break Before (hrs)",
                "Long Shift", "Short Break", "Fatigue Risk"
            ]
            disp["Long Shift"]   = disp["Long Shift"].map({True: "⚠️ YES", False: ""})
            disp["Short Break"]  = disp["Short Break"].map({True: "⚠️ YES", False: ""})
            disp["Fatigue Risk"] = disp["Fatigue Risk"].map({True: "🔴 YES", False: ""})
            st.dataframe(disp, use_container_width=True, hide_index=True)

    with tab2:
        wdf = r["weekly_df"].copy()
        wdf["Status"] = wdf["Exceeds 60h"].map({True: "🔴 EXCEEDS 60h", False: "✅ OK"})
        wdf = wdf.drop(columns=["Exceeds 60h"])
        st.dataframe(wdf, use_container_width=True, hide_index=True)

    # Download button
    st.download_button(
        f"⬇️ Download Report — {stem}.xlsx",
        data=r["excel_bytes"],
        file_name=f"Compliance_{stem}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"dl_{stem}",
    )
    st.divider()
